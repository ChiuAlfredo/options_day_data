import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
import os

# 讀取Excel文件
df = pd.read_excel("data.xlsx")

# 將 '時間' 列轉換為datetime格式，只提取時間部分
df['Time'] = pd.to_datetime(df['時間'], errors='coerce').dt.time

# 定義時間範圍
start_time = pd.to_datetime('09:00:00').time()
end_time = pd.to_datetime('09:10:00').time()

# 篩選時間範圍
timefiltered_df = df[(df['Time'] >= start_time) & (df['Time'] <= end_time)]

# 篩選 '價性' & '類型'
ks="深度價外 (DOTM)"
option_C="C"
option_P="P"
ksdf_C = timefiltered_df[(timefiltered_df['價性'] == ks) & (timefiltered_df['買賣權'] == option_C)].copy()
ksdf_P = timefiltered_df[(timefiltered_df['價性'] == ks) & (timefiltered_df['買賣權'] == option_P)].copy()

# 計算 '選擇權成交價變化率' 和 '期貨價格變化率' 的相關係數並四捨五入到小數第四位，略過空值
correlationTXF_C = ksdf_C['選擇權成交價變化率'].corr(ksdf_C['期貨價格變化率'])
roundedTXF_C = round(correlationTXF_C, 4)
correlationTXF_P = ksdf_P['選擇權成交價變化率'].corr(ksdf_P['期貨價格變化率'])
roundedTXF_P = round(correlationTXF_P, 4)

# 計算 '選擇權成交價變化率' 和 '指數價格變化率' 的相關係數並四捨五入到小數第四位，略過空值
correlationTW_C = ksdf_C['選擇權成交價變化率'].corr(ksdf_C['指數價格變化率'])
roundedTW_C = round(correlationTW_C, 4)
correlationTW_P = ksdf_P['選擇權成交價變化率'].corr(ksdf_P['指數價格變化率'])
roundedTW_P = round(correlationTW_P, 4)

# 計算 '期貨價格變化率' 和 '指數價格變化率' 的相關係數並四捨五入到小數第四位，略過空值
correlationTXFTW = timefiltered_df['期貨價格變化率'].corr(timefiltered_df['指數價格變化率'])
roundedTXFTW = round(correlationTXFTW, 4)

# 計算資料筆數
rows_C = ksdf_C['選擇權成交價變化率'].notna().sum()-1
rows_P = ksdf_P['選擇權成交價變化率'].notna().sum()-1
rows_TXFTW = timefiltered_df['指數價格變化率'].notna().sum()-1

# 顯示相關係數 & 資料數
print(ks,option_C)
print(f"{roundedTXF_C}")
print(f"{roundedTW_C}")
print(f"{rows_C}")
print(ks,option_P)
print(f"{roundedTXF_P}")
print(f"{roundedTW_P}")
print(f"{rows_P}")
print("加權期貨")
print(f"{roundedTXFTW}")
print(f"{rows_TXFTW}")
