#我把filtered_df裡的9.的資料拿掉了 print出來的filtered_df_without9沒有問題 但grouped_df裡還是有9.的值
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
import os

# 讀取Excel文件
df = pd.read_excel("data.xlsx")

# 假設 '時間' 列包含日期和時間，將其轉換為datetime格式
df['Datetime'] = pd.to_datetime(df['時間'], errors='coerce')

# 提取日期和時間部分
df['Date'] = df['Datetime'].dt.date
df['Time'] = df['Datetime'].dt.time

# 定義時間範圍
start_time = pd.to_datetime('09:00:00').time()
end_time = pd.to_datetime('13:30:00').time()

# 篩選時間範圍，並排除09:00:00的數據
filtered_df = df[(df['Time'] > start_time) & (df['Time'] <= end_time)]
filtered_df_without9 = filtered_df[filtered_df['Datetime'] != pd.to_datetime('2021-01-06 09:00:00')]

# 按履約價和每5分鐘分組計算總和
grouped_df = filtered_df_without9.groupby(['履約價', '買賣權', pd.Grouper(key='Datetime', freq='5min')]).sum().reset_index()

# 選擇所需的列並打印
selected_columns = grouped_df[['履約價', '買賣權', 'Datetime', '指數價格變化率']]
print(selected_columns)
