import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
import numpy as np

class Record:
    def __init__(self, result_queue,size):
        self.record_list= self._ques_to_list(result_queue,size)
        self.profit_list = self._profit_record()
        self.open_list = self._position_record('open')
        self.close_list = self._position_record('close')
        self.earn_list,self.loss_list = self._earn_loss_record()
        self.statistic_dict = self._statistic()
    

    def _ques_to_list(self, result_queue,size):
         # 按顺序获取结果
        record_list = [None] * size
        while not result_queue.empty():
            index, record_dict = result_queue.get()
            record_list[index] = record_dict

        # 移除 None 值
        record_list = [record for record in record_list if record is not None]
        
        return record_list
    
    def _profit_record(self):
        profit_list = []
        for index,record in enumerate(self.record_list):
            profit = 0
            # record = record_list[1]
            if record['期貨紀錄'] != []:
                profit +=record['期貨紀錄'][-1]['損益累計']
            if record['選擇權紀錄'] != []:
                profit +=record['選擇權紀錄'][-1]['損益累計']

            profit_list.append(float(profit))
        
        return profit_list
    
    def _position_record(self,position_status):
        position_status_list = []
        if position_status=='open':
            position_status_chines = '開倉'
        else:
            position_status_chines = '平倉'
            
        for index,record in enumerate(self.record_list):
            open = 0
            open_future_sum = 0
            open_option_sum = 0
            open_future = []
            open_option = []
            # record = record_list[1]
            if record['期貨紀錄'] != []:
                open_future = [i['損益'] for i in record['期貨紀錄'] if i['倉位狀況'] == position_status]
                open_future_sum = sum(open_future)
                
            if record['選擇權紀錄'] != []:
                open_option = [ i['損益']  for i in record['選擇權紀錄'] if i['倉位狀況']==position_status]
                open_option_sum = sum(open_option)
              
            

            open = open_future_sum + open_option_sum
            position_status_list.append({
                f'{position_status_chines}損益':open,
                f'{position_status_chines}期貨損益':open_future_sum,
                f'{position_status_chines}期貨個別損益':open_future,
                f'{position_status_chines}選擇權損益':open_option_sum,
                f'{position_status_chines}選擇權個別損益':open_option
            })
            
        return position_status_list
    
    def _earn_loss_record(self):
        earn_list = [profit for profit in self.profit_list if profit > 0]
        loss_list = [profit for profit in self.profit_list if profit <= 0]
        return earn_list,loss_list
        
    
    def _statistic(self):
        period = f'{self.record_list[0]["日期"]}~{self.record_list[-1]["日期"]}'
        total_profit = sum(self.earn_list)+sum(self.loss_list)
        total_earn = sum(self.earn_list)
        total_loss = sum(self.loss_list)
        total_trade_days = len(self.profit_list)
        
        win_trade_count = len(self.earn_list)
        loss_trade_count = len(self.profit_list)
        
        win_rate = len(self.earn_list)/len(self.profit_list)
  
        max_earn_trade = max(self.profit_list)
        min_loss_trade = min(self.profit_list)
        
        avg_earn_trade = total_earn/win_trade_count
        avg_loss_trade = total_loss/loss_trade_count
        earn_loss_rate = avg_earn_trade/avg_loss_trade
        avg_trade_profit = sum(self.profit_list)/len(self.profit_list)
        std_trade_profit = np.std(self.profit_list)
        
        avg_open_cost = sum([i['開倉損益'] for i in  self.open_list])/total_trade_days
        
        
        
        
        stdescribe_dict = {
            '期間':period,
            '總淨利':total_profit,
            '毛利':total_earn,
            '毛損失':total_loss,
            '總交易天數':total_trade_days,
            '勝率':win_rate,
            '成功筆數':win_trade_count,
            '失敗筆數':loss_trade_count,
            '最大獲利交易':max_earn_trade,
            '最大虧損交易':min_loss_trade,
            '成功交易平均獲利':avg_earn_trade,
            '失敗交易平均虧損':avg_loss_trade,
            '平均獲利/平均虧損':earn_loss_rate,
            '平均每筆交易盈虧':avg_trade_profit,
            '每筆交易盈虧標準差':std_trade_profit,
            '平均開倉成本':avg_open_cost,
            '平均報酬率':avg_trade_profit/avg_open_cost,
            
        }
        return stdescribe_dict
    
    def show_statistic(self):
        return self.statistic_dict
    
    def output_excel(self,name):
        with pd.ExcelWriter(f'{name}.xlsx') as writer:
            self._ouput_st_describe(writer)
            self._output_profit(writer)
            self._ouput_record(writer)
        self._add_daily_profit_chart(name)
        self._add_accumulate_profit_chart(name)
        print("檔案已儲存為",f'{name}.xlsx')

            
           

        print("record_list 已保存為 record_list.xlsx")
    def _ouput_st_describe(self,writer):
        statistic_df = pd.DataFrame(self.statistic_dict, index=[0])
        statistic_df.transpose().to_excel(writer, sheet_name='統計', header=False)
        
    def _output_profit(self,writer):
        date_list = [i["日期"] for i in self.record_list]
        ac_profit_list = np.cumsum(self.profit_list)
        combined_list = [
            {
                '日期': date,
                '損益': profit,
                '累積損益': ac_profit,
                **{f'{k}': v for k, v in open_data.items()},
                **{f'{k}': v for k, v in close_data.items()}
            }
            for date, profit,ac_profit, open_data, close_data in zip(date_list, self.profit_list,ac_profit_list, self.open_list, self.close_list)
        ]
                
        profit_df = pd.DataFrame(combined_list)
        profit_df.to_excel(writer, sheet_name='每日損益', index=False)
        
        

            
        
    def _ouput_record(self,writer):
        for record in self.record_list:
            # record = record_list[0]
            # 將每個記錄轉換為 DataFrame
            df = pd.DataFrame()
            if record['期貨紀錄'] == []:
                df_future = pd.DataFrame(columns=['時間_期貨紀錄','收盤價_期貨紀錄','數量_期貨紀錄','手續費_期貨紀錄','損益_期貨紀錄','損益累計_期貨紀錄'])
            else:
                df_future = pd.DataFrame(record['期貨紀錄']).add_suffix('_期貨紀錄')
            
            if record['選擇權紀錄'] == []:
                df_option = pd.DataFrame(columns=['時間_選擇權紀錄','BorS_選擇權紀錄','倉位狀況_選擇權紀錄','履約價_選擇權紀錄','CorP_選擇權紀錄','價格_選擇權紀錄','數量_選擇權紀錄','手續費_選擇權紀錄','損益_選擇權紀錄','損益累計_選擇權紀錄'])
            else:
                df_option = pd.DataFrame(record['選擇權紀錄']).add_suffix('_選擇權紀錄')
            
            if 'call_價格' in record:
                record['call_價格'] = record['call_價格'].add_suffix('_call')
            else:
                record['call_價格'] = pd.DataFrame(columns=['時間_call','價格_call'])
            
            if 'put_價格' in record:
                record['put_價格'] = record['put_價格'].add_suffix('_put')
            else:
                record['put_價格'] = pd.DataFrame(columns=['時間_put','價格_put'])
            
            df = pd.merge(record['加權指數'],record['期貨'], left_on='分鐘時間',right_on='時間', how='inner', suffixes=('_加權指數', '_期貨'))
            df = df[['時間','收盤價_加權指數','收盤價_期貨']]
            df = pd.merge(df,df_option, left_on='時間',right_on='時間_選擇權紀錄', how='left', suffixes=('', '_選擇權紀錄'))
            df = pd.merge(df,df_future, left_on='時間',right_on='時間_期貨紀錄', how='left', suffixes=('', '_期貨紀錄'))
            
            df = pd.merge(df,record['call_價格'], left_on='時間',right_on='時間_call', how='left', suffixes=('', '_call'))
            df = pd.merge(df,record['put_價格'], left_on='時間',right_on='時間_put', how='left', suffixes=('', '_put'))
            
            df.sort_values(by='時間', inplace=True)
            
            # 使用日期作為分頁名稱
            sheet_name = record['日期']
            
            
            # 將 DataFrame 寫入 Excel 分頁
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _add_daily_profit_chart(self,name):
        wb = load_workbook(f'{name}.xlsx')
        ws = wb['每日損益']

        chart = LineChart()
        chart.title = "每日損益折線圖"
        chart.style = 13
        chart.y_axis.title = '損益'
        chart.x_axis.title = '日期'

        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=ws.max_row)

        chart.add_data(data, titles_from_data=True)


        ws.add_chart(chart, "Q2")

        wb.save(f'{name}.xlsx')
    def _add_accumulate_profit_chart(self,name):
        wb = load_workbook(f'{name}.xlsx')
        ws = wb['每日損益']

        chart = LineChart()
        chart.title = "每日累積損益折線圖"
        chart.style = 13
        chart.y_axis.title = '累積損益'
        chart.x_axis.title = '日期'

        data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=ws.max_row)

        chart.add_data(data, titles_from_data=True)


        ws.add_chart(chart, "Z2")

        wb.save(f'{name}.xlsx')
    
    
        
        
    
    